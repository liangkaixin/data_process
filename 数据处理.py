# /Users/imperial.liang/Documents/数据处理/2024年度举报投诉数据_原始数据.xlsx
import logging
from datetime import datetime

import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from constant import PROVINCES, CITY_TO_PROVINCE, COUNTY_TO_PROVINCE, KEYWORDS, MARKET_CAT_DICT

# 配置日志
logging.basicConfig(
    filename=f"举报数据处理_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log",
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console.setFormatter(formatter)
logger.addHandler(console)

# 文件路径
# 用户输入文件名
file_path = input("请输入原始数据文件名（带扩展名）: ")
output_file = '整体.xlsx'
# 读取数据
try:
    logger.info(f"开始读取文件: {file_path}")
    df_query = pd.read_excel(file_path, sheet_name='举报查询')
    df_closed = pd.read_excel(file_path, sheet_name='外部已办结举报')
    df_government = pd.read_excel(file_path, sheet_name='全国交办举报')
    logger.info(f"成功读取文件: {file_path}")
except FileNotFoundError as e:
    logger.error(f"文件不存在: {e}")
    exit(1)
except Exception as e:
    logger.error(f"读取文件时出错: {e}")
    exit(1)

# 模板列
template_columns = [
    '渠道', '季度', '交办地区', '举报标题', '月份', '举报日期', '举报内容', '举报人姓名', '举报人性别',
    '举报人手机号', '举报人通讯地址', '被举报单位名称', '被举报单位联系方式', '被举报单位通讯地址',
    '举报方式', '办理结果', '投诉事由', '举报投诉信息办理地区区划', '举报人邮箱',
    '被举报人所属区划', '举报投诉类型', '举报分类', '区域区划', '来源', '举报类别二类']

# 映射字段
query_map = {
    '举报标题': '举报标题',
    '举报日期': '举报日期',
    '举报内容': '举报内容',
    '举报人姓名': '举报人姓名',
    '性别': '举报人性别',
    '联系电话': '举报人手机号',
    '通讯地址': '举报人通讯地址',
    '被举报单位名称': '被举报单位名称',
    '被举报单位联系方式': '被举报单位联系方式',
    '被举报单位通讯地址': '被举报单位通讯地址',
    '举报方式': '举报方式',
    '举报分类': '举报分类',
    '举报区域区划': '区域区划',
    '回复内容': '办理结果',
    '举报分类': '举报类别二类'
}

closed_map = {
    '举报标题': '举报标题',
    '举报日期': '举报日期',
    '举报内容': '举报内容',
    '举报人姓名': '举报人姓名',
    '举报人性别': '举报人性别',
    '举报人手机号': '举报人手机号',
    '举报人通讯地址': '举报人通讯地址',
    '被举报单位名称': '被举报单位名称',
    '被举报单位联系方式': '被举报单位联系方式',
    '被举报单位通讯地址': '被举报单位通讯地址',
    '举报方式': '举报方式',
    '办理结果': '办理结果',
    '举报投诉信息办理地区区划': '举报投诉信息办理地区区划',
    '举报人邮箱': '举报人邮箱',
    '被举报人所属区划': '被举报人所属区划',
    '举报投诉类型': '举报投诉类型',
    '举报分类': '举报分类',
    '区域区划': '区域区划',
    '投诉事由': '投诉事由',
    '来源': '来源',
    '举报类别': '举报类别二类'
}

government_map = {'性别': '举报人性别',
                  '通讯地址': '举报人通讯地址',
                  '举报投诉状态': '举报状态',
                  '联系电话': '举报人手机号',
                  '举报类别': '举报类别二类'
                  }

# 获取模板列名
logger.info(f"获取模板列名，共{len(template_columns)}列")


# 标准化并补齐字段
def standardize_df(df, mapping):
    logger.info(f"开始标准化DataFrame，原始形状: {df.shape}")
    df_std = df.rename(columns=mapping)
    for col in template_columns:
        if col not in df_std.columns:
            df_std[col] = pd.NA
    logger.info(f"完成标准化DataFrame，标准化后形状: {df_std.shape}")
    return df_std[template_columns]


# 应用标准化
logger.info("开始处理'举报查询'数据")
df_query_std = standardize_df(df_query, query_map)
logger.info("开始处理'外部已办结举报'数据")
df_closed_std = standardize_df(df_closed, closed_map)
logger.info("开始处理'全国交办举报'数据")
df_government = df_government.rename(columns=government_map)

# 添加渠道标识
df_query_std['渠道'] = '举报查询'
df_closed_std['渠道'] = '外部已办结'
logger.info("成功添加渠道标识")

# 合并数据
logger.info("开始合并数据")
df_merged = pd.concat([df_query_std, df_closed_std], ignore_index=True)
logger.info(f"数据合并完成，合并后形状: {df_merged.shape}")

# 输出预览
print(df_merged.head())
logger.info("输出数据前5行预览")

# 处理缺失值
logger.info("开始处理缺失值")
df_merged['举报投诉类型'] = df_merged['举报投诉类型'].fillna('举报')

missing_counts = df_merged.isna().sum()
for col, count in missing_counts.items():
    if count > 0:
        logger.info(f"列 '{col}' 有 {count} 个缺失值")
logger.info("缺失值处理完成")

# 指定用于去重的列
dedup_columns = ['举报标题', '投诉事由', '举报人姓名', '举报人性别', '举报人手机号', '被举报单位名称']
logger.info(f"指定去重列: {dedup_columns}")

# 初步去重
logger.info("开始基于指定列去重")
df_deduplicated = df_merged.drop_duplicates(subset=dedup_columns, keep='first')
df_government = df_government.drop_duplicates(
    subset=['举报标题', '举报人姓名', '举报人性别', '举报人手机号', '被举报单位名称'], keep='first')
logger.info(f"去重完成，去重前记录数: {len(df_merged)}, 去重后记录数: {len(df_deduplicated)}")

# 基于内容相似度的高级去重
df = df_deduplicated.copy()
logger.info("开始基于内容相似度去重")

# 过滤掉关键信息缺失的记录
df = df[df['举报人姓名'].notna() & df['举报人手机号'].notna()].copy()
logger.info(f"过滤关键信息缺失的记录，过滤前记录数: {len(df_deduplicated)}, 过滤后记录数: {len(df)}")

# 新建保留标记列
df['保留标记'] = False

# 按举报人分组
grouped = df.groupby(['举报人姓名', '举报人手机号'])
logger.info(f"按举报人姓名和手机号分组，共{len(grouped)}组")

final_rows = []
total_groups = len(grouped)

for group_num, (_, group) in enumerate(grouped, 1):
    if group_num % 100 == 0:
        logger.info(f"处理分组 {group_num}/{total_groups}")

    used = [False] * len(group)
    indices = group.index.tolist()

    for i in range(len(group)):
        if used[i]:
            continue

        ref_idx = indices[i]
        ref_title = str(group.loc[ref_idx, '举报标题'])
        ref_content = str(group.loc[ref_idx, '举报内容'])

        similar_group = [ref_idx]
        used[i] = True

        # 查找相似的举报
        for j in range(i + 1, len(group)):
            if used[j]:
                continue

            cmp_idx = indices[j]
            cmp_title = str(group.loc[cmp_idx, '举报标题'])
            cmp_content = str(group.loc[cmp_idx, '举报内容'])

            # 计算相似度
            sim_title = fuzz.token_set_ratio(ref_title, cmp_title)
            sim_content = fuzz.token_set_ratio(ref_content, cmp_content)

            # 标题和内容均高相似度才算重复
            if sim_title > 90 and sim_content > 90:
                similar_group.append(cmp_idx)
                used[j] = True

        # 选择要保留的记录
        sub_df = df.loc[similar_group]

        if '举报状态' in sub_df.columns:
            if '已办结' in sub_df['举报状态'].values:
                # 优先保留已办结的记录
                chosen = sub_df[sub_df['举报状态'] == '已办结'].iloc[0]
            else:
                # 保留内容最长的记录
                chosen = sub_df.loc[sub_df['举报内容'].str.len().idxmax()]
        else:
            # 保留内容最长的记录
            chosen = sub_df.loc[sub_df['举报内容'].str.len().idxmax()]

        final_rows.append(chosen)

# 构建最终去重后的DataFrame
df_deduplicated = pd.DataFrame(final_rows).drop(columns='保留标记', errors='ignore')
logger.info(f"内容相似度去重完成，去重前记录数: {len(df)}, 去重后记录数: {len(df_deduplicated)}")

# 结果查看
print(f"原始数量：{len(df)}, 去重后数量：{len(df_deduplicated)}")
logger.info(f"最终去重结果：原始数量：{len(df)}, 去重后数量：{len(df_deduplicated)}")
df = df_deduplicated.copy()


# 从文本中提取省级名称
def extract_province(text):
    if pd.isna(text) or text == '':
        return ''

    # 首先尝试直接匹配省份
    for province in PROVINCES:
        if province in str(text):
            return province

    # 尝试通过城市名查找省份
    for city, province in CITY_TO_PROVINCE.items():
        if city in str(text):
            return province

    # 尝试通过区名查找省份
    for county, province in COUNTY_TO_PROVINCE.items():
        if county in str(text) and county != '':
            return province

    return ''


# 按优先级从多个字段中查找省级名称
def get_province(row):
    for field in df.columns:
        if field in row:
            province = extract_province(row[field])
            if province:
                return province
    return ''


def dateTransfer(df, tpye='举报'):
    # 确保投诉日期列是datetime类型
    df[f'{tpye}日期'] = pd.to_datetime(df[f'{tpye}日期'], format='mixed')

    # 添加季度字段 (格式: 2025Q1)
    df['季度'] = df[f'{tpye}日期'].dt.year.astype(str) + 'Q' + df[f'{tpye}日期'].dt.quarter.astype(str)

    # 添加月度字段 (格式: 2025年3月)
    df['月度'] = df[f'{tpye}日期'].dt.year.astype(str) + '年' + df[f'{tpye}日期'].dt.month.astype(str) + '月'

    # 调整列的顺序，确保月度放在季度后面
    df = df[[f'{tpye}日期', '季度', '月度'] + [col for col in df.columns if col not in [f'{tpye}日期', '季度', '月度']]]
    return df


# 应用函数创建新字段
logger.info("开始提取交办地区信息")
df['交办地区'] = df.apply(get_province, axis=1)
df_government['交办地区'] = df_government.apply(get_province, axis=1)
province_counts = df['交办地区'].value_counts()
for province, count in province_counts.items():
    if province:
        logger.info(f"地区 '{province}' 有 {count} 条记录")
logger.info("交办地区信息提取完成")

# 确保举报日期列是datetime类型
df = dateTransfer(df, '举报')
df_government = dateTransfer(df_government, '举报')
logger.info("日期字段处理完成")

# 删除举报状态中包含不予受理的数据
logger.info("开始过滤包含'不予受理'的记录")
original_count = len(df)
df = df[~df.apply(lambda row: row.astype(str).str.contains('不予受理', na=False).any(), axis=1)]
filtered_count = len(df)
logger.info(
    f"过滤完成，过滤前记录数: {original_count}, 过滤后记录数: {filtered_count}, 共过滤掉 {original_count - filtered_count} 条记录")

# # 增加对应市场和所属大类字段
code_cat_market_dic = pd.DataFrame({'二类编码':
                                        [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21,
                                         22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35],
                                    '举报类别二类':
                                        [
                                            'qxz', '互联网上网服务营业场所', '歌舞娱乐场所', '游艺娱乐场所',
                                            '其他娱乐场所', '艺术品经营单位', '互联网文化单位', '网络游戏经营单位',
                                            '电影发行放映场所', '广播电台电视台', '互联网视听节目服务单位',
                                            '音像制品经营单位', '印刷企业', '互联网出版机构', '出版物经营单位',
                                            '电影发行放映场所', '其他', '文化表演团体', '演出经纪机构',
                                            '演出场所经营单位', '个体演员', '个体演出经纪人', '文物经营单位',
                                            '文物保护单位', '文物收藏单位', '卫星传送的境外电视节目接收单位',
                                            '其他娱乐场所', '社会艺术水平考级机构', '旅行社', '旅行社分社',
                                            '旅行社服务网点', '导游', '旅行社相关人员', '其他旅游经营活动',
                                            '在线旅游经营者', '网络表演经纪机构'],
                                    '对应市场':
                                        ['旅游', '文化', '文化', '文化', '文化', '文化', '文化', '出版', '电影',
                                         '广播电视', '广播电视', '出版', '出版', '出版', '出版', '电影', '其他', '文化',
                                         '文化', '文化', '文化', '文化', '文物', '文物', '文物', '广播电视', '文化',
                                         '文化', '旅游', '旅游', '旅游', '旅游', '旅游', '旅游', '旅游', '文化'],
                                    '大类':
                                        ['旅游', '文化', '文化', '文化', '文化', '文化', '文化', '其他', '其他', '其他',
                                         '其他', '其他', '其他', '其他', '其他', '其他', '其他', '文化', '文化', '文化',
                                         '文化', '文化', '其他', '其他', '其他', '其他', '文化', '文化', '旅游', '旅游',
                                         '旅游', '旅游', '旅游', '旅游', '旅游', '文化',
                                         ]}
                                   )


#
# # 应用模糊匹配函数
# logger.info("开始基于字典关联举报类别二类，得到对应市场和大类")
# df = df.merge(code_cat_market_dic, on = '举报类别二类', how='left')
# df_government = df_government.merge(code_cat_market_dic, on = '举报类别二类', how='left')
# logger.info("基于字典关联举报类别二类，完成")

# 模糊匹配函数，增加对应市场和所属大类字段
def match_keywords(content, title, keyword_dict, origin_cat):
    for market, market_dict in keyword_dict.items():
        for category, keywords_dict in market_dict.items():
            for report_type, keywords_list in keywords_dict.items():
                # 合并举报类型和关键词列表
                keywords = keywords_list + [report_type]

                # 先匹配举报标题
                if any(fuzz.partial_ratio(keyword, title) > 80 for keyword in keywords):
                    return MARKET_CAT_DICT.get(market), market, category

                # 再匹配举报内容
                if any(fuzz.partial_ratio(keyword, content) > 80 for keyword in keywords):
                    return MARKET_CAT_DICT.get(market), market, category
    for key, value in keyword_dict.items():
        if str(origin_cat) in str(value):
            return MARKET_CAT_DICT.get(key), key, origin_cat


# 应用模糊匹配函数
logger.info("开始基于关键词模糊匹配补充'对应市场'和'所属大类'")
df[['大类', '对应市场', '举报类别二类']] = df.apply(
    lambda x: pd.Series(match_keywords(x['举报内容'], x['举报标题'], KEYWORDS, x['举报类别二类'])),
    axis=1
)
logger.info("关键词模糊匹配完成")

# 根据对应市场中文化和其他市场分类，删除包含特定关键词的行
logger.info("开始第一次关键词过滤")
keywords_to_delete1 = ['消费纠纷', '会员卡', '充值', '充钱', '游戏币']
mask_culture_other = df['对应市场'].isin(['文化', '其他'])
mask_keywords1 = df.apply(
    lambda row: any(
        keyword in str(row['举报标题']) or keyword in str(row['举报内容']) for keyword in keywords_to_delete1),
    axis=1
)
original_count = len(df)
df = df.drop(df[mask_culture_other & mask_keywords1].index)
filtered_count = len(df)
logger.info(
    f"第一次关键词过滤完成，过滤前记录数: {original_count}, 过滤后记录数: {filtered_count}, 共过滤掉 {original_count - filtered_count} 条记录")

# 根据对应市场中其他-文化-电影发行放映场所，删除包含特定关键词的行
logger.info("开始第二次关键词过滤")
keywords_to_delete2 = ['消费纠纷', '会员卡', '充值', '充钱', '退票', '退费', '退款']
mask_movie = df['举报类别二类'].str.contains('电影发行放映场所', na=False)
mask_keywords2 = df.apply(
    lambda row: any(
        keyword in str(row['举报标题']) or keyword in str(row['举报内容']) for keyword in keywords_to_delete2),
    axis=1
)
original_count = len(df)
df = df.drop(df[mask_movie & mask_keywords2].index)
filtered_count = len(df)
logger.info(
    f"第二次关键词过滤完成，过滤前记录数: {original_count}, 过滤后记录数: {filtered_count}, 共过滤掉 {original_count - filtered_count} 条记录")

# 保存结果
logger.info("开始保存结果文件")
try:
    df.to_excel(output_file, sheet_name='举报整体', index=False)
    logger.info(f"数据已成功保存到 '总体.xlsx'，最终记录数: {len(df)}")
    book = load_workbook(output_file)
    # 创建新的 sheet
    sheet = book.create_sheet('全国交办举报')
    # 使用 pandas 将 DataFrame 写入新的 sheet
    for r in dataframe_to_rows(df_government, index=False, header=True):
        sheet.append(r)
    # 保存文件
    book.save(output_file)
except Exception as e:
    logger.error(f"保存文件时出错: {e}")

try:
    df_query = pd.read_excel(file_path, sheet_name='投诉查询')
    df_closed = pd.read_excel(file_path, sheet_name='外部已办结投诉')
    df_government = pd.read_excel(file_path, sheet_name='全国交办投诉')

    logger.info(f"投诉查询表: {len(df_query)} 行")
    logger.info(f"外部已办结投诉表: {len(df_closed)} 行")

    # 模板列
    template_columns = [
        '渠道', '季度', '办理地区', '投诉标题', '月份', '投诉日期', '投诉内容', '投诉人姓名', '投诉人性别',
        '投诉人手机号', '投诉人通讯地址', '被投诉单位名称', '被投诉单位联系方式', '被投诉单位通讯地址',
        '旅游类别', '旅游方式', '目的地', '受理人数', '投诉方式', '投诉对象', '投诉问题', '投诉原因', '办理结果',
        '文件名', '服务质量问题发生地', '国籍', '客源地', '证件类型', '合同日期', '投诉信息办理地区区划',
        '举报投诉类型', '客源地区划', '服务质量问题发生地区划', '目的地区划', '投诉请求事项', '投诉事由', '来源'
    ]

    # 映射字段
    query_map = {
        '投诉标题': '投诉标题',
        '投诉日期': '投诉日期',
        '投诉问题': '投诉内容',
        '投诉人姓名': '投诉人姓名',
        '性别': '投诉人性别',
        '联系电话': '投诉人手机号',
        '通讯地址': '投诉人通讯地址',
        '被投诉单位名称': '被投诉单位名称',
        '被投诉单位联系方式': '被投诉单位联系方式',
        '被投诉单位通讯地址': '被投诉单位通讯地址',
        '投诉方式': '投诉方式',
        '投诉分类': '投诉分类',
        '投诉区域区划': '区域区划',
        '回复内容': '办理结果',
        '投诉事由': '投诉事由',
        '客源地': '客源地',
        '证件类型': '证件类型',
        '合同日期': '合同日期',
        '旅游类别': '旅游类别',
        '旅游方式': '旅游方式',
        '目的地': '目的地',
        '受理人数': '受理人数',
        '投诉请求事项': '投诉请求事项',
        '来源': '来源'

    }

    closed_map = {
        '投诉标题': '投诉标题',
        '投诉日期': '投诉日期',
        '投诉内容': '投诉内容',
        '投诉人姓名': '投诉人姓名',
        '投诉人性别': '投诉人性别',
        '投诉人手机号': '投诉人手机号',
        '投诉人通讯地址': '投诉人通讯地址',
        '被投诉单位名称': '被投诉单位名称',
        '被投诉单位联系方式': '被投诉单位联系方式',
        '被投诉单位通讯地址': '被投诉单位通讯地址',
        '投诉方式': '投诉方式',
        '办理结果': '办理结果',
        '投诉信息办理地区区划': '投诉信息办理地区区划',
        '投诉人邮箱': '投诉人邮箱',
        '被投诉人所属区划': '被投诉人所属区划',
        '举报投诉类型': '举报投诉类型',
        '投诉分类': '投诉分类',
        '区域区划': '区域区划',
        '投诉事由': '投诉事由',
        '来源': '来源'
    }

    government_map = {'性别': '投诉人性别',
                      '通讯地址': '投诉人通讯地址',
                      '举报投诉状态': '投诉状态'
                      }
    logger.info(f"模板列数: {len(template_columns)}")
    logger.info(f"模板列名: {', '.join(template_columns)}")


    # 标准化并补齐字段
    def standardize_df(df, mapping):
        df_std = df.rename(columns=mapping)
        for col in template_columns:
            if col not in df_std.columns:
                df_std[col] = None  # 或 pd.NA
        return df_std[template_columns]


    # 应用
    logger.info("开始标准化数据")
    df_query_std = standardize_df(df_query, query_map)
    df_closed_std = standardize_df(df_closed, closed_map)
    logger.info(f"投诉查询标准化后: {len(df_query_std)} 行")
    logger.info(f"外部已办结投诉标准化后: {len(df_closed_std)} 行")

    df_government = standardize_df(df_government, closed_map)

    df_query_std['渠道'] = '投诉查询'
    df_closed_std['渠道'] = '外部已办结投诉'

    # 合并
    logger.info("开始合并数据")
    df_merged = pd.concat([df_query_std, df_closed_std], ignore_index=True)
    logger.info(f"合并后数据: {len(df_merged)} 行")

    df_merged = pd.concat([df_query_std, df_closed_std], ignore_index=True)

    df_merged['对应市场'] = '旅游'
    df_merged['大类'] = '旅游'
    df_government['对应市场'] = '旅游'
    df_government['大类'] = '旅游'
    logger.info("完成新增 对应市场 和 大类， 值都是 旅游")

    df_merged['举报投诉类型'] = df_merged['举报投诉类型'].fillna('投诉')
    logger.info(f"举报投诉类型分布: {df_merged['举报投诉类型'].value_counts().to_dict()}")

    # 假设df是你的DataFrame
    df = df_merged
    # 先把需要用到的字段都转换成字符串，防止空值或非字符串类型影响拼接
    # 根据投诉标题+投诉内容+举报人+手机号+被投诉单位去重
    for col in ['投诉标题', '投诉内容', '投诉人姓名', '举报投诉类型', '投诉人手机号', '被投诉单位名称', '投诉事由']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str)
        else:
            df[col] = ''  # 防止列缺失


    # 定义一个拼接字段，作为去重的依据
    def build_key(row):
        if row['投诉内容'] == '' or pd.isna(row['投诉内容']):
            # 投诉内容为空时
            key = (
                    row['投诉标题'] + row['投诉人姓名'] +
                    row['投诉人手机号'] + row['被投诉单位名称']
            )
        else:
            # 投诉内容不为空时
            key = (
                    row['投诉标题'] + row['投诉人姓名'] +
                    row['投诉人手机号'] + row['被投诉单位名称'] + row['投诉内容']
            )
        return key


    logger.info("开始基于拼接键去重")
    df['去重键'] = df.apply(build_key, axis=1)

    # 根据去重键去重
    df_deduplicated = df.drop_duplicates(subset=['去重键']).drop(columns=['去重键'])
    logger.info(f"基于拼接键去重后: {len(df_deduplicated)} 行，减少了 {len(df) - len(df_deduplicated)} 行")

    # 假设 df 是已标准化后的 DataFrame
    df = df_deduplicated.copy()

    # 分离投诉内容为空的行（它们不参与去重，但要保留）
    df_empty = df[df['投诉内容'].isna() | (df['投诉内容'].astype(str).str.strip() == '')].copy()
    logger.info(f"投诉内容为空的记录: {len(df_empty)} 行")

    # 过滤掉投诉内容为空的行，进入去重逻辑
    df = df[~df.index.isin(df_empty.index)]
    logger.info(f"参与相似度去重的记录: {len(df)} 行")

    # 再过滤掉投诉人姓名或手机号为空的（相同判断前提）
    df = df[df['投诉人姓名'].notna() & df['投诉人手机号'].notna()]
    logger.info(f"投诉人姓名和手机号都不为空的记录: {len(df)} 行")

    # 初始化标记列
    df['保留标记'] = False

    # 分组去重
    logger.info("开始基于相似度去重")
    grouped = df.groupby(['投诉人姓名', '投诉人手机号'])
    final_rows = []
    total_groups = len(grouped)
    processed_groups = 0

    for _, group in grouped:
        processed_groups += 1
        if processed_groups % 100 == 0:
            logger.info(f"处理相似度分组: {processed_groups}/{total_groups}")

        used = [False] * len(group)
        indices = group.index.tolist()

        for i in range(len(group)):
            if used[i]:
                continue

            ref_idx = indices[i]
            ref_title = str(group.loc[ref_idx, '投诉标题']) if pd.notna(group.loc[ref_idx, '投诉标题']) else ''
            ref_content = str(group.loc[ref_idx, '投诉内容']) if pd.notna(group.loc[ref_idx, '投诉内容']) else ''
            if not ref_content.strip():
                continue

            similar_group = [ref_idx]
            used[i] = True

            for j in range(i + 1, len(group)):
                if used[j]:
                    continue

                cmp_idx = indices[j]
                cmp_title = str(group.loc[cmp_idx, '投诉标题']) if pd.notna(group.loc[cmp_idx, '投诉标题']) else ''
                cmp_content = str(group.loc[cmp_idx, '投诉内容']) if pd.notna(group.loc[cmp_idx, '投诉内容']) else ''
                if not cmp_content.strip():
                    continue

                sim_title = fuzz.token_set_ratio(ref_title, cmp_title)
                sim_content = fuzz.token_set_ratio(ref_content, cmp_content)

                if sim_title > 90 and sim_content > 90:
                    similar_group.append(cmp_idx)
                    used[j] = True

            # 在 similar_group 中选择保留哪一个
            sub_df = df.loc[similar_group]

            if '投诉状态' in sub_df.columns:
                if '已办结' in sub_df['投诉状态'].values:
                    chosen = sub_df[sub_df['投诉状态'] == '已办结'].iloc[0]
                else:
                    chosen = sub_df.loc[sub_df['投诉内容'].str.len().idxmax()]
            else:
                chosen = sub_df.loc[sub_df['投诉内容'].str.len().idxmax()]

            final_rows.append(chosen)

    # 拼接最终结果：去重后数据 + 原始“投诉内容为空”的数据
    df_deduplicated = pd.DataFrame(final_rows)
    df_deduplicated = pd.concat([df_deduplicated, df_empty], ignore_index=True).drop(columns='保留标记',
                                                                                     errors='ignore')
    logger.info(
        f"相似度去重后: {len(df_deduplicated)} 行，原始数量: {len(df) + len(df_empty)}，减少了 {len(df) + len(df_empty) - len(df_deduplicated)} 行")

    df = df_deduplicated

    # 补充“交办地区”，举报查询中已有交办地区，外部已办结中部分地区为空，需要根据表中“质量发生地”进行人工填入
    logger.info("开始补充交办地区")


    # 创建函数从文本中提取省级名称
    def extract_province(text):
        if pd.isna(text) or text == '':
            return ''

        # 首先尝试直接匹配省份
        for province in PROVINCES:
            if province in str(text):
                return province

        # 如果没有直接匹配到省份，尝试通过城市名查找省份
        for city, province in CITY_TO_PROVINCE.items():
            if city in str(text):
                return province

        # 如果没有直接匹配到省份，尝试通过区名查找省份
        for county, province in COUNTY_TO_PROVINCE.items():
            if county in str(text) and county != '':
                return province

        return ''


    # 按优先级从多个字段中查找省级名称
    def get_province(row):
        # 检查各个字段，按优先级顺序
        for field in df.columns:
            if field in row:
                province = extract_province(row[field])
                if province:
                    return province
        return ''


    # 应用函数创建新字段
    df['交办地区'] = df.apply(get_province, axis=1)
    df_government['交办地区'] = df_government.apply(get_province, axis=1)
    logger.info(f"补充交办地区完成，成功补充: {df['交办地区'].count()} 条记录")

    #  补充“季度”“月度”，根据投诉、投诉日期，补充季度、月度，方便报告筛选和执法局后续数据提取。
    logger.info("开始补充季度和月度信息")

    df = dateTransfer(df, '投诉')
    df_government = dateTransfer(df_government, '投诉')
    logger.info(f"季度分布: {df['季度'].value_counts().to_dict()}")
    logger.info(f"月度分布: {df['月度'].value_counts().to_dict()}")

    # 删除举报状态中包含不予受理的数据
    logger.info("开始过滤包含'不予受理'的记录")
    original_count = len(df)
    df = df[~df.apply(lambda row: row.astype(str).str.contains('不予受理', na=False).any(), axis=1)]
    filtered_count = len(df)
    logger.info(
        f"过滤完成，过滤前记录数: {original_count}, 过滤后记录数: {filtered_count}, 共过滤掉 {original_count - filtered_count} 条记录")

    # 标记在线投诉旅行社
    logger.info("开始标记在线投诉旅行社")
    ota_keywords = [
        'ota', '线上', '抖音', '携程网', 'ctrip', '微信群', '飞猪', '去哪儿', '趣拿',
        '美团', '直播', '快手', '短视频', '携程', '同程', '驴妈妈'
    ]

    # 将关键词小写化（用于不区分大小写匹配）
    ota_keywords_lower = [kw.lower() for kw in ota_keywords]

    # 条件1：投诉对象 == '投诉旅行社'
    cond_base = df['投诉对象'] == '投诉旅行社'


    # 条件2：从所有列中匹配关键词 —— 推荐从具体文本字段中找，比如“投诉内容”或“来源”等

    def contains_ota_in_any_column(row):
        for col in df.columns:  # 遍历DataFrame的所有列
            text = row[col]
            if pd.notna(text):
                text = str(text).lower()
                # print(text)  # 调试用打印语句，实际使用时可注释掉
                if any(kw in text for kw in ota_keywords_lower):
                    return True
        return False


    # 应用于整个DataFrame的每一行
    cond_ota = df.apply(contains_ota_in_any_column, axis=1)

    # 同时满足“投诉对象 == 投诉旅行社” 且包含 OTA 关键词的行，标记为 “在线投诉旅行社”
    original_travel_agency_count = len(df[df['投诉对象'] == '投诉旅行社'])
    df.loc[cond_base & cond_ota, '投诉对象'] = '在线投诉旅行社'
    new_online_travel_agency_count = len(df[df['投诉对象'] == '在线投诉旅行社'])
    logger.info(f"标记在线投诉旅行社完成，共标记: {new_online_travel_agency_count} 条记录")

    # 加载现有的工作簿
    book = load_workbook(output_file)

    # 创建新的 sheet '投诉整体' 并将 df 数据写入
    sheet = book.create_sheet('投诉整体')
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # 创建新的 sheet '全国交办投诉' 并将 df_government 数据写入
    sheet = book.create_sheet('全国交办投诉')
    for r in dataframe_to_rows(df_government, index=False, header=True):
        sheet.append(r)
    book.save(output_file)

    logger.info(f"数据处理完成，结果已保存到: {output_file}，最终记录数: {len(df)}")
except Exception as e:
    logger.error(f"处理过程中发生错误: {str(e)}", exc_info=True)
